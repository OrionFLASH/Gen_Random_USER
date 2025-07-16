import os
import csv
import random
import string
import pandas as pd
import logging
from datetime import datetime, timedelta
from collections import Counter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from dateutil.relativedelta import relativedelta  # для агрегатора
import time



MODE = "AGGREGATE"   # "GENERATE" или "AGGREGATE"

SUMMARY_PARAMS = {
    "summary_sheet": "SUMMARY",
    "summary_csv_base": "SUMMARY",  # без расширения
    "fields": ["TN", "LastName", "FirstName", "MidleName", "TB_CODE", "ROLE_CODE"],  # что брать из штатки
    "dupe_col": "IsDuplicate",
    "month_periods": [f"M-{i}" for i in range(13)],
    "week_periods": [f"N-{i}" for i in range(9)],
    "2week_periods": [f"2N-{i}" for i in range(3)],
}


AGG_PARAMS = {
    "input_dir": "//Users//orionflash//Desktop//MyProject//Gen_Random_USER//IN_BASE",
    "input_base": "EMPLOYEE_VISITS_400",   # Без расширения
    "input_role_base": "EMPLOEE_400",      # Без расширения
    "output_dir": "//Users//orionflash//Desktop//MyProject//Gen_Random_USER//OUT_CSV",
    "output_base": "VISIT_AGGREGATED_400",                 # Без расширения
    "log_dir": "//Users//orionflash//Desktop//MyProject//Gen_Random_USER//LOGS",
    "log_base": "AGG_LOG",
    "period_day_month_border": 7,  # до какого числа месяца включительно M-0 объединяет два месяца
    "period_weekday_border": 1, # до какого дня недели включительно N-0 и 2N-0 захватывают прошлую неделю/две (0=Пн, 1=Вт и т.д.)
    # Ожидаемые поля
    "field_tn": "TN",
    "field_lastname": "LastName",
    "field_firstname": "FirstName",
    "field_middlename": "MidleName",
    "field_tb": "TB_CODE",
    "field_date": "Date",
    "field_nvisits": "Visited",
    "field_role": "ROLE_CODE"
}


# ======= ПАРАМЕТРЫ ==========
PARAMS = {
    "input_dir": "//Users//orionflash//Desktop//MyProject//Gen_Random_USER//IN_BASE",
    "org_unit_file": "ORG_UNIT.csv",
    "output_dir": "//Users//orionflash//Desktop//MyProject//Gen_Random_USER//OUT_CSV",
    "output_base": "EMPLOEE",    # Без расширения
    "log_dir": "//Users//orionflash//Desktop//MyProject//Gen_Random_USER//LOGS",
    "log_base": "LOG",
    "user_count": 8000,
    "role_distribution": {
        "KM_KKSB": 1700,
        "KM_MNS": 200,
        "RUK_KKSB": 320,
        "RUK_MNS": 15,
        "RUK_TB": 30,
        "RUK_CA": 10,
        "DRKB": 50,
        "KM_SB1": 2000,
        "RUK_SB1": 200,
        "UMB1": 500,
        "RUK_UMB1": 10,
        "OTHER_CA": 30,
        "SERVISE_MAN": 1700
    }
}
# ============================

VISIT_PARAMS = {
    "visit_out_dir": "//Users//orionflash//Desktop//MyProject//Gen_Random_USER//OUT_CSV",
    "visit_out_base": "EMPLOYEE_VISITS",    # Без расширения
    "visit_start_date": "2025-01-01",
    "visit_end_date":   "2025-08-05",
    "visit_target_row_count": 400_000
}

# --- Системные текстовки для итогового лога ---
SUMMARY_TEXT = {
    "time_start": "Время старта программы: {ts}",
    "time_end": "Время окончания программы: {ts}",
    "time_total": "Общее время работы: {sec:.2f} сек",
    "block": "----- {block} -----",
    "block_time": "Блок '{block}': {sec:.2f} сек",
    "gen_users": "Сгенерировано сотрудников: {count}",
    "gen_visits": "Сгенерировано посещений: {count}",
    "agg_visits": "Обработано посещений: {count}",
    "agg_users": "Обработано уникальных сотрудников: {count}",
    "mean_role": "Среднее число посещений по роли '{role}': {mean:.2f}",
}


def get_timestamp(fmt="%Y%m%d_%H%M%S"):
    return datetime.now().strftime(fmt)

def get_log_file():
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"{PARAMS['log_base']}_{date_str}.log"
    return os.path.join(PARAMS['log_dir'], filename)

def setup_logging(log_dir, log_base):
    os.makedirs(log_dir, exist_ok=True)
    logger = logging.getLogger()
    logger.handlers.clear()
    logger.setLevel(logging.INFO)
    # Файл
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"{log_base}_{date_str}.log"
    log_path = os.path.join(log_dir, filename)
    file_handler = logging.FileHandler(log_path, encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    # Консоль
    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(logging.INFO)
    # Формат логов
    fmt = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s')
    file_handler.setFormatter(fmt)
    stream_handler.setFormatter(fmt)
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    logger.info("=== START RUN ===")
    return logger


def log_time(logger, block, t_start, t_end, text_dict=SUMMARY_TEXT):
    sec = t_end - t_start
    logger.info(text_dict["block_time"].format(block=block, sec=sec))

def log_summary(logger, summary: dict, text_dict=SUMMARY_TEXT):
    for key, val in summary.items():
        logger.info(val)


def autofit_excel_columns(xlsx_path, sheet_name="Sheet1"):
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    # Заголовки жирные
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # Автоширина колонок
    for column_cells in ws.columns:
        max_len = 0
        col = column_cells[0].column_letter
        for cell in column_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except:
                pass
        adjusted_width = max_len + 2
        ws.column_dimensions[col].width = adjusted_width
    wb.save(xlsx_path)

def load_org_units(path, logger):
    try:
        df = pd.read_csv(path, sep=';', dtype=str)
        units = df[['TB_CODE','TB_FULL_NAME','TB_SHORT_NAME','GOSB_CODE','GOSB_NAME','GOSB_SHORT_NAME','ORG_UNIT_CODE']].to_dict('records')
        logger.info(f"Загружено подразделений: {len(units)} из {path}")
        return units
    except Exception as e:
        logger.error(f"Ошибка при загрузке оргструктуры: {e}")
        raise


# --- ФИО ---
LAST_NAMES_MALE = [
    "Иванов", "Петров", "Сидоров", "Кузнецов", "Попов", "Васильев", "Смирнов", "Михайлов", "Новиков", "Федоров",
    "Белов", "Семенов", "Егоров", "Козлов", "Соловьев", "Калинин", "Тихонов", "Жуков", "Орлов", "Макаров", "Чернов"
]
LAST_NAMES_FEMALE = [x + "а" if not x.endswith("в") else x[:-1] + "ва" for x in LAST_NAMES_MALE]
FIRST_NAMES_MALE = [
    "Алексей", "Иван", "Дмитрий", "Сергей", "Виктор", "Владимир", "Павел", "Егор", "Константин", "Георгий",
    "Максим", "Андрей", "Артем", "Никита", "Роман", "Олег", "Ярослав", "Денис", "Станислав", "Виталий", "Юрий"
]
FIRST_NAMES_FEMALE = [
    "Елена", "Мария", "Анна", "Ирина", "Ольга", "Наталья", "Светлана", "Татьяна", "Алина", "Екатерина",
    "Дарья", "Юлия", "Кристина", "Марина", "Валерия", "Виктория", "Полина", "Вероника", "Евгения", "София", "Алёна"
]
MIDDLE_NAMES_MALE = [
    "Алексеевич", "Иванович", "Дмитриевич", "Сергеевич", "Викторович", "Владимирович", "Павлович", "Егорович",
    "Константинович", "Георгиевич", "Максимович", "Андреевич", "Артемович", "Никитович", "Романович", "Олегович",
    "Ярославович", "Денисович", "Станиславович", "Витальевич", "Юрьевич"
]
MIDDLE_NAMES_FEMALE = [x[:-2] + "вна" for x in MIDDLE_NAMES_MALE]

def generate_unique_fio_set(count, logger):
    fio_set = set()
    results = []
    gender_list = []
    male_count = count // 2
    female_count = count - male_count

    logger.info(f"Планируется мужчин: {male_count}, женщин: {female_count}")

    # Генерируем мужчин
    for _ in range(male_count):
        tries = 0
        while True:
            ln = random.choice(LAST_NAMES_MALE)
            fn = random.choice(FIRST_NAMES_MALE)
            mn = random.choice(MIDDLE_NAMES_MALE)
            fio = f"{ln} {fn} {mn}"
            if fio not in fio_set:
                fio_set.add(fio)
                results.append((ln, fn, mn))
                gender_list.append("M")
                break
            tries += 1
            if tries > 1000:
                logger.warning("Проблема с уникальностью мужских ФИО, расширьте список.")
                break

    # Генерируем женщин
    for _ in range(female_count):
        tries = 0
        while True:
            ln = random.choice(LAST_NAMES_FEMALE)
            fn = random.choice(FIRST_NAMES_FEMALE)
            mn = random.choice(MIDDLE_NAMES_FEMALE)
            fio = f"{ln} {fn} {mn}"
            if fio not in fio_set:
                fio_set.add(fio)
                results.append((ln, fn, mn))
                gender_list.append("F")
                break
            tries += 1
            if tries > 1000:
                logger.warning("Проблема с уникальностью женских ФИО, расширьте список.")
                break

    # Если сотрудников больше, чем вариантов, будем добавлять рандомные буквы
    while len(results) < count:
        gender = random.choice(["M", "F"])
        if gender == "M":
            ln = random.choice(LAST_NAMES_MALE) + random.choice(string.ascii_uppercase)
            fn = random.choice(FIRST_NAMES_MALE) + random.choice(string.ascii_uppercase)
            mn = random.choice(MIDDLE_NAMES_MALE)
        else:
            ln = random.choice(LAST_NAMES_FEMALE) + random.choice(string.ascii_uppercase)
            fn = random.choice(FIRST_NAMES_FEMALE) + random.choice(string.ascii_uppercase)
            mn = random.choice(MIDDLE_NAMES_FEMALE)
        fio = f"{ln} {fn} {mn}"
        if fio not in fio_set:
            fio_set.add(fio)
            results.append((ln, fn, mn))
            gender_list.append(gender)
    logger.info(f"Итого уникальных ФИО: {len(results)} (мужчин: {gender_list.count('M')}, женщин: {gender_list.count('F')})")
    return results, gender_list


def generate_unique_tn_set(count, logger):
    numbers = set()
    while len(numbers) < count:
        num = str(random.randint(10**3, 10**10-1)).zfill(10)
        numbers.add(num)
    logger.info(f"Итого уникальных табельных номеров: {len(numbers)}")
    return list(numbers)


def distribute_users(units, n_users, logger):
    n_units = len(units)
    base = n_users // n_units
    result = []
    rest = n_users
    for i, u in enumerate(units):
        if i < n_units - 2:
            cnt = base + random.randint(-int(0.2*base), int(0.2*base))
        else:
            cnt = random.randint(1,2)
        if cnt > rest:
            cnt = rest
        result.extend([i]*cnt)
        rest -= cnt
    for i in range(rest):
        result.append(i % n_units)
    random.shuffle(result)
    logger.info(f"Распределено сотрудников по подразделениям: {Counter(result)}")
    return result


def generate_employees(org_units, n_users, role_distribution, logger):
    fio_list, gender_list = generate_unique_fio_set(n_users, logger)
    tn_list = generate_unique_tn_set(n_users, logger)
    unit_indices = distribute_users(org_units, n_users, logger)

    roles = []
    total_roles = sum(role_distribution.values())
    for role, cnt in role_distribution.items():
        roles.extend([role]*cnt)
    if len(roles) < n_users:
        roles.extend(["OTHER"]*(n_users-len(roles)))
    elif len(roles) > n_users:
        roles = roles[:n_users]
    random.shuffle(roles)

    employees = []
    for idx in range(n_users):
        org = org_units[unit_indices[idx]]
        ln, fn, mn = fio_list[idx]
        tn = tn_list[idx]
        role = roles[idx]
        gender = gender_list[idx]
        emp = {
            "TN": tn,
            "LastName": ln,
            "FirstName": fn,
            "MidleName": mn,
            "Gender": gender,
            "ROLE_CODE": role,
            "TB_CODE": org["TB_CODE"],
            "TB_FULL_NAME": org["TB_FULL_NAME"],
            "TB_SHORT_NAME": org["TB_SHORT_NAME"],
            "GOSB_CODE": org["GOSB_CODE"],
            "GOSB_NAME": org["GOSB_NAME"],
            "GOSB_SHORT_NAME": org["GOSB_SHORT_NAME"],
            "ORG_UNIT_CODE": org["ORG_UNIT_CODE"]
        }
        employees.append(emp)
    logger.info(f"Распределено по ролям: {Counter(roles)}")
    return employees



def generate_visits(employees_df, start_date, end_date, target_row_count, logger):
    dates = []
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    delta = end_dt - start_dt
    dates = [start_dt + timedelta(days=i) for i in range(delta.days+1)]

    emp_count = len(employees_df)
    visit_data = []
    emp_indices = employees_df.index.tolist()
    random.shuffle(emp_indices)

    rare_indices = set(emp_indices[:int(emp_count*0.3)])
    active_indices = set(emp_indices[int(emp_count*0.3):int(emp_count*0.35)])
    usual_indices = set(emp_indices) - rare_indices - active_indices

    all_weekdays = [d for d in dates if d.weekday() < 5]
    all_weekends = [d for d in dates if d.weekday() >= 5]

    # Основной цикл по сотрудникам
    for idx in emp_indices:
        emp = employees_df.loc[idx]
        if idx in rare_indices:
            visit_days = random.randint(1, 5)
            visit_dates = random.sample(dates, visit_days)
        elif idx in active_indices:
            visit_dates = set(all_weekdays)
            extra_weekends = random.sample(all_weekends, min(len(all_weekends), random.randint(5, 20)))
            visit_dates = list(visit_dates.union(extra_weekends))
            skip_cnt = random.randint(0, 5)
            if skip_cnt:
                visit_dates = random.sample(visit_dates, len(visit_dates) - skip_cnt)
        else:
            n_days = random.randint(10, 50)
            candidate_dates = all_weekdays + random.sample(all_weekends, min(len(all_weekends), random.randint(0, 3)))
            visit_dates = random.sample(candidate_dates, min(n_days, len(candidate_dates)))

        for dt in visit_dates:
            row = dict(emp)
            row['Date'] = dt.strftime("%Y-%m-%d")
            row['Visited'] = 1
            visit_data.append(row)
        if len(visit_data) >= target_row_count:
            break

    # Добор случайных визитов если строк меньше чем нужно
    if len(visit_data) < target_row_count:
        logger.warning(f"Недостаточно визитов: {len(visit_data)}. Добавляем случайные визиты до {target_row_count}...")
        all_dates = [dt.strftime("%Y-%m-%d") for dt in dates]
        emp_indices_full = employees_df.index.tolist()
        while len(visit_data) < target_row_count:
            idx = random.choice(emp_indices_full)
            emp = employees_df.loc[idx]
            dt = random.choice(all_dates)
            row = dict(emp)
            row['Date'] = dt
            row['Visited'] = 1
            visit_data.append(row)

    # Обрезаем если вдруг перебор
    if len(visit_data) > target_row_count:
        visit_data = visit_data[:target_row_count]
    logger.info(f"Сгенерировано посещений: {len(visit_data)}")
    return visit_data


def build_month_periods(max_date, n_periods, border_day, logger=None):
    """
    Возвращает список кортежей (start, end) для каждого периода M-0, M-1, ... (без перекрытия!)
    """
    periods = []
    if max_date.day <= border_day:
        # M-0 объединяет два месяца: предыдущий полностью + текущий до max_date
        m0_start = (max_date - pd.DateOffset(months=1)).replace(day=1)
        m0_end = max_date
        periods.append((m0_start, m0_end))
        if logger:
            logger.info(f"M-0: {m0_start.date()} .. {m0_end.date()} (Объединённый период)")
        # Следующий период: месяц, предшествующий m0_start
        for i in range(1, n_periods):
            # start: первый день месяца на i месяцев до m0_start
            start = (m0_start - pd.DateOffset(months=i)).replace(day=1)
            end = (start + pd.DateOffset(months=1)) - pd.DateOffset(days=1)
            periods.append((start, end))
            if logger:
                logger.info(f"M-{i}: {start.date()} .. {end.date()}")
    else:
        # Обычный режим: M-0 только последний месяц
        m0_start = max_date.replace(day=1)
        m0_end = max_date
        periods.append((m0_start, m0_end))
        if logger:
            logger.info(f"M-0: {m0_start.date()} .. {m0_end.date()}")
        for i in range(1, n_periods):
            start = (m0_start - pd.DateOffset(months=i)).replace(day=1)
            end = (start + pd.DateOffset(months=1)) - pd.DateOffset(days=1)
            periods.append((start, end))
            if logger:
                logger.info(f"M-{i}: {start.date()} .. {end.date()}")
    return periods



def build_week_periods(max_date, n_periods, border_weekday, logger=None):
    """
    Возвращает список (start, end) для каждой недели: N-0, N-1... без перекрытия!
    N-0 может быть объединённым (две недели), остальные — обычные недели.
    border_weekday: до какого дня недели включительно захватываем прошлую неделю (0=Пн)
    """
    periods = []
    max_weekday = max_date.weekday()
    if max_weekday <= border_weekday:
        # N-0: прошлый понедельник - max_date (охватывает почти две недели)
        n0_start = max_date - timedelta(days=max_weekday + 7)
        n0_end = max_date
        periods.append((n0_start, n0_end))
        if logger:
            logger.info(f"N-0: {n0_start.date()} .. {n0_end.date()} (Объединённый период)")
        # N-1: предыдущая неделя до n0_start-1
        for i in range(1, n_periods):
            # новый период — неделя ДО предыдущего периода
            start = n0_start - timedelta(weeks=i)
            end = start + timedelta(days=6)
            periods.append((start, end))
            if logger:
                logger.info(f"N-{i}: {start.date()} .. {end.date()}")
    else:
        # N-0: текущий понедельник - max_date
        n0_start = max_date - timedelta(days=max_weekday)
        n0_end = max_date
        periods.append((n0_start, n0_end))
        if logger:
            logger.info(f"N-0: {n0_start.date()} .. {n0_end.date()}")
        for i in range(1, n_periods):
            start = n0_start - timedelta(weeks=i)
            end = start + timedelta(days=6)
            periods.append((start, end))
            if logger:
                logger.info(f"N-{i}: {start.date()} .. {end.date()}")
    return periods



def get_period_label(dt, periods, label_template, default=None):
    """
    Присваивает метку периоду по номеру (находит первое попадание dt в период).
    label_template: "M-{}", "N-{}", "2N-{}"
    default: что вернуть, если не найдено совпадений (по умолчанию — последняя метка)
    """
    for idx, (start, end) in enumerate(periods):
        if start <= dt <= end:
            return label_template.format(idx)
    if default is not None:
        return default
    return label_template.format(len(periods)-1)



def build_2week_periods(max_date, n_periods, border_weekday, logger=None):
    """
    Возвращает список (start, end) для каждого двухнедельного периода: 2N-0, 2N-1...
    2N-0 может быть объединённым (3 недели), остальные — обычные двухнедельки.
    """
    periods = []
    max_weekday = max_date.weekday()
    if max_weekday <= border_weekday:
        # 2N-0: прошлый-прошлый понедельник - max_date (охватывает почти 3 недели)
        n0_start = max_date - timedelta(days=max_weekday + 14)
        n0_end = max_date
        periods.append((n0_start, n0_end))
        if logger:
            logger.info(f"2N-0: {n0_start.date()} .. {n0_end.date()} (Объединённый период)")
        for i in range(1, n_periods):
            start = n0_start - timedelta(weeks=2*i)
            end = start + timedelta(days=13)
            periods.append((start, end))
            if logger:
                logger.info(f"2N-{i}: {start.date()} .. {end.date()}")
    else:
        # 2N-0: текущий понедельник - max_date (две недели)
        n0_start = max_date - timedelta(days=max_weekday)
        n0_end = max_date
        periods.append((n0_start, n0_end))
        if logger:
            logger.info(f"2N-0: {n0_start.date()} .. {n0_end.date()}")
        for i in range(1, n_periods):
            start = n0_start - timedelta(weeks=2*i)
            end = start + timedelta(days=13)
            periods.append((start, end))
            if logger:
                logger.info(f"2N-{i}: {start.date()} .. {end.date()}")
    return periods



def save_employees(employees, out_dir, out_base, logger, timestamp=None):
    os.makedirs(out_dir, exist_ok=True)
    ts = timestamp or get_timestamp()
    csv_name = f"{out_base}_{ts}.csv"
    xlsx_name = f"{out_base}_{ts}.xlsx"
    csv_path = os.path.join(out_dir, csv_name)
    xlsx_path = os.path.join(out_dir, xlsx_name)

    fields = list(employees[0].keys())
    df = pd.DataFrame(employees)
    df.to_csv(csv_path, sep=";", index=False)
    logger.info(f"Данные сотрудников сохранены в CSV: {csv_path}")

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="GEN_USER", index=False)
    autofit_excel_columns(xlsx_path, sheet_name="GEN_USER")
    logger.info(f"Данные сотрудников сохранены в Excel: {xlsx_path}")

    return df, csv_path, xlsx_path




def save_visits(visit_data, out_csv, out_xlsx, logger):
    df = pd.DataFrame(visit_data)
    df.to_csv(out_csv, sep=";", index=False)
    logger.info(f"Таблица посещений сохранена в CSV: {out_csv}")

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="VISITS", index=False)
    autofit_excel_columns(out_xlsx, sheet_name="VISITS")
    logger.info(f"Таблица посещений сохранена в Excel: {out_xlsx}")

def generate_all(logger):
    prog_t0 = time.time()
    logger.info("----- ГЕНЕРАЦИЯ -----")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    org_path = os.path.join(PARAMS["input_dir"], PARAMS["org_unit_file"])

    t0 = time.time()
    org_units = load_org_units(org_path, logger)
    t1 = time.time()
    logger.info(f"Время на загрузку оргструктуры: {t1-t0:.2f} сек")

    t2 = time.time()
    employees = generate_employees(org_units, PARAMS["user_count"], PARAMS["role_distribution"], logger)
    t3 = time.time()
    logger.info(f"Время на генерацию сотрудников: {t3-t2:.2f} сек")
    logger.info(f"Сгенерировано сотрудников: {len(employees)}")

    t4 = time.time()
    employees_df, _, _ = save_employees(
        employees,
        PARAMS["output_dir"],
        PARAMS["output_base"],
        logger,
        timestamp=timestamp
    )
    t5 = time.time()
    logger.info(f"Время на сохранение сотрудников: {t5-t4:.2f} сек")

    logger.info("Генерация сотрудников завершена, начинаем генерацию посещений")

    visit_csv = os.path.join(VISIT_PARAMS["visit_out_dir"], f"{VISIT_PARAMS['visit_out_base']}_{timestamp}.csv")
    visit_xlsx = os.path.join(VISIT_PARAMS["visit_out_dir"], f"{VISIT_PARAMS['visit_out_base']}_{timestamp}.xlsx")

    t6 = time.time()
    visit_data = generate_visits(
        employees_df,
        VISIT_PARAMS['visit_start_date'],
        VISIT_PARAMS['visit_end_date'],
        VISIT_PARAMS['visit_target_row_count'],
        logger
    )
    t7 = time.time()
    logger.info(f"Время на генерацию посещений: {t7-t6:.2f} сек")

    save_visits(visit_data, visit_csv, visit_xlsx, logger)
    t8 = time.time()
    logger.info(f"Время на сохранение посещений: {t8-t7:.2f} сек")
    logger.info("Генерация посещений завершена")

    # Лог финальных метрик
    df_visits = pd.DataFrame(visit_data)
    role_means = df_visits.groupby("ROLE_CODE")["Visited"].mean().to_dict() if not df_visits.empty else {}
    logger.info(f"ИТОГ: сгенерировано сотрудников: {len(employees)}, посещений: {len(visit_data)}")
    for role, mean in role_means.items():
        logger.info(f"Среднее число посещений по роли '{role}': {mean:.2f}")
    prog_t1 = time.time()
    logger.info(f"Время выполнения generate_all: {prog_t1-prog_t0:.2f} сек")


def aggregate(logger):
    prog_t0 = time.time()
    logger.info("----- АГРЕГАЦИЯ -----")
    params = AGG_PARAMS
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Параметры для периодов
    period_day_month_border = params.get("period_day_month_border", 7)
    period_weekday_border = params.get("period_weekday_border", 1)

    input_csv = os.path.join(params["input_dir"], params["input_base"] + ".csv")
    input_role_csv = os.path.join(params["input_dir"], params["input_role_base"] + ".csv")
    output_csv = os.path.join(params["output_dir"], f"{params['output_base']}_{timestamp}.csv")
    output_xlsx = os.path.join(params["output_dir"], f"{params['output_base']}_{timestamp}.xlsx")

    t0 = time.time()
    logger.info("Режим агрегации посещений")
    df = pd.read_csv(input_csv, sep=";", dtype=str)
    df[params["field_nvisits"]] = pd.to_numeric(df[params["field_nvisits"]], errors="coerce").fillna(0).astype(int)
    df_role = pd.read_csv(input_role_csv, sep=";", dtype=str)
    t1 = time.time()
    logger.info(f"Время на загрузку файлов: {t1-t0:.2f} сек")

    key_fields = [
        params["field_tn"], params["field_lastname"],
        params["field_firstname"], params["field_middlename"],
        params["field_tb"], params["field_date"]
    ]
    agg_dict = {params["field_nvisits"]: "sum"}
    logger.info(f"Агрегируем по: {key_fields}, агрегируем поле: {params['field_nvisits']}")
    t2 = time.time()
    df_agg = df.groupby(key_fields, as_index=False).agg(agg_dict)
    t3 = time.time()
    logger.info(f"Время на агрегацию данных: {t3-t2:.2f} сек")

    # --- ДОБАВЛЯЕМ КОЛОНКИ МЕСЯЦ, НЕДЕЛЯ, ДВУХНЕДЕЛЬНЫЙ ПЕРИОД ---
    logger.info("Определяем максимальную дату для расчёта периодов...")
    max_date = pd.to_datetime(df_agg[params["field_date"]]).max()

    n_months = len(SUMMARY_PARAMS["month_periods"])
    n_weeks = len(SUMMARY_PARAMS["week_periods"])
    n_2weeks = len(SUMMARY_PARAMS["2week_periods"])

    month_periods = build_month_periods(max_date, n_months, period_day_month_border, logger)
    week_periods = build_week_periods(max_date, n_weeks, period_weekday_border, logger)
    twoweek_periods = build_2week_periods(max_date, n_2weeks, period_weekday_border, logger)

    t4 = time.time()
    logger.info("Добавляем временные колонки (месяц, неделя, двухнедельный период)...")
    df_agg["DATE_DT"] = pd.to_datetime(df_agg[params["field_date"]])

    df_agg["PERIOD_MONTH"] = df_agg["DATE_DT"].apply(lambda dt: get_period_label(dt, month_periods, "M-{}"))
    df_agg["PERIOD_WEEK"] = df_agg["DATE_DT"].apply(lambda dt: get_period_label(dt, week_periods, "N-{}"))
    df_agg["PERIOD_2WEEK"] = df_agg["DATE_DT"].apply(lambda dt: get_period_label(dt, twoweek_periods, "2N-{}"))
    t5 = time.time()
    logger.info(f"Время на добавление временных колонок: {t5-t4:.2f} сек")

    # --- ДОБАВЛЯЕМ ДАННЫЕ ПО РОЛИ ---
    logger.info("Дозаполнение ROLE_CODE из штатки (join по TN)...")
    df_role_part = df_role[[params["field_tn"], params["field_role"]]].drop_duplicates()
    df_result = df_agg.merge(df_role_part, how="left", on=params["field_tn"])
    t6 = time.time()
    logger.info(f"Время на join с ROLE_CODE: {t6-t5:.2f} сек")
    logger.info(f"Финальный размер: {df_result.shape}")

    # === СОЗДАНИЕ SUMMARY-ЛИСТА ===
    tn_counts = df_role[params["field_tn"]].value_counts()
    df_role["IsDuplicate"] = df_role[params["field_tn"]].map(lambda x: "Дубль" if tn_counts[x] > 1 else "")

    summary_fields = SUMMARY_PARAMS["fields"] + [SUMMARY_PARAMS["dupe_col"]]
    df_summary = df_role[summary_fields].drop_duplicates()

    agg = df_agg.copy()
    agg["DATE_DT"] = pd.to_datetime(agg[params["field_date"]])

    # Месяцы
    for i, m_label in enumerate(SUMMARY_PARAMS["month_periods"]):
        start, end = month_periods[i]
        mask = (agg["DATE_DT"] >= start) & (agg["DATE_DT"] <= end)
        visited = agg.loc[mask].groupby(params["field_tn"])[params["field_nvisits"]].sum().gt(0)
        df_summary[m_label] = df_summary["TN"].map(lambda tn: int(visited.get(tn, False)))
    # Недели
    for i, w_label in enumerate(SUMMARY_PARAMS["week_periods"]):
        start, end = week_periods[i]
        mask = (agg["DATE_DT"] >= start) & (agg["DATE_DT"] <= end)
        visited = agg.loc[mask].groupby(params["field_tn"])[params["field_nvisits"]].sum().gt(0)
        df_summary[w_label] = df_summary["TN"].map(lambda tn: int(visited.get(tn, False)))
    # 2-недельные
    for i, tw_label in enumerate(SUMMARY_PARAMS["2week_periods"]):
        start, end = twoweek_periods[i]
        mask = (agg["DATE_DT"] >= start) & (agg["DATE_DT"] <= end)
        visited = agg.loc[mask].groupby(params["field_tn"])[params["field_nvisits"]].sum().gt(0)
        df_summary[tw_label] = df_summary["TN"].map(lambda tn: int(visited.get(tn, False)))

    t7 = time.time()
    logger.info(f"Время на подготовку SUMMARY: {t7-t6:.2f} сек")

    summary_csv = os.path.join(params["output_dir"], f"{SUMMARY_PARAMS['summary_csv_base']}_{timestamp}.csv")
    df_summary.to_csv(summary_csv, sep=";", index=False)
    logger.info(f"Summary сохранён в CSV: {summary_csv}")

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        df_result.to_excel(writer, sheet_name="AGGREGATED", index=False)
        df_summary.to_excel(writer, sheet_name=SUMMARY_PARAMS["summary_sheet"], index=False)
    autofit_excel_columns(output_xlsx, sheet_name="AGGREGATED")
    autofit_excel_columns(output_xlsx, sheet_name=SUMMARY_PARAMS["summary_sheet"])
    logger.info(f"Результат сохранён в Excel: {output_xlsx} (автоформатирование)")

    df_result.to_csv(output_csv, sep=";", index=False)
    logger.info(f"Результат сохранён в CSV: {output_csv}")

    # Финальная статистика по ролям
    role_means = df_result.groupby("ROLE_CODE")[params["field_nvisits"]].mean().to_dict() if not df_result.empty else {}
    logger.info(f"ИТОГ: обработано посещений: {len(df)}, уникальных записей: {len(df_result)}")
    for role, mean in role_means.items():
        logger.info(f"Среднее число посещений по роли '{role}': {mean:.2f}")

    prog_t1 = time.time()
    logger.info(f"Время выполнения aggregate: {prog_t1-prog_t0:.2f} сек")





def create_summary_sheet(logger, df_role, df_agg, max_date, output_dir, summary_csv_base):
    logger.info("Создаём summary-лист по всем сотрудникам и периодам...")

    # Считаем дубли по TN
    tn_counts = df_role["TN"].value_counts()
    df_role["IsDuplicate"] = df_role["TN"].map(lambda x: "Дубль" if tn_counts[x] > 1 else "")

    # Оставляем только нужные колонки и дубли
    summary_fields = SUMMARY_PARAMS["fields"] + [SUMMARY_PARAMS["dupe_col"]]
    df_summary = df_role[summary_fields].drop_duplicates()

    # Для ускорения — индексы по TN
    agg = df_agg.copy()
    agg["DATE_DT"] = pd.to_datetime(agg["Date"])

    # Периоды
    # Месяцы
    for i, m_label in enumerate(SUMMARY_PARAMS["month_periods"]):
        start = (max_date - pd.DateOffset(months=i)).replace(day=1)
        end = (start + pd.DateOffset(months=1)) - pd.DateOffset(days=1)
        mask = (agg["DATE_DT"] >= start) & (agg["DATE_DT"] <= end)
        visited = agg.loc[mask].groupby("TN")["Visited"].sum().gt(0)
        df_summary[m_label] = df_summary["TN"].map(lambda tn: int(visited.get(tn, False)))
    # Недели
    for i, w_label in enumerate(SUMMARY_PARAMS["week_periods"]):
        week_start = (max_date - pd.DateOffset(weeks=i)).to_period('W').start_time
        week_end = week_start + pd.Timedelta(days=6)
        mask = (agg["DATE_DT"] >= week_start) & (agg["DATE_DT"] <= week_end)
        visited = agg.loc[mask].groupby("TN")["Visited"].sum().gt(0)
        df_summary[w_label] = df_summary["TN"].map(lambda tn: int(visited.get(tn, False)))
    # 2-недельные
    for i, tw_label in enumerate(SUMMARY_PARAMS["2week_periods"]):
        tw_start = (max_date - pd.DateOffset(weeks=2*i)).to_period('W').start_time
        tw_end = tw_start + pd.Timedelta(days=13)
        mask = (agg["DATE_DT"] >= tw_start) & (agg["DATE_DT"] <= tw_end)
        visited = agg.loc[mask].groupby("TN")["Visited"].sum().gt(0)
        df_summary[tw_label] = df_summary["TN"].map(lambda tn: int(visited.get(tn, False)))

    # Сохраняем в CSV
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    summary_csv = os.path.join(output_dir, f"{summary_csv_base}_{timestamp}.csv")
    df_summary.to_csv(summary_csv, sep=";", index=False)
    logger.info(f"Summary сохранён в CSV: {summary_csv}")

    return df_summary



def main():
    prog_start = time.time()
    if MODE.upper() == "AGGREGATE":
        logger = setup_logging(AGG_PARAMS["log_dir"], AGG_PARAMS["log_base"])
        logger.info("Старт режима АГРЕГАЦИЯ")
        aggregate(logger)
    elif MODE.upper() == "GENERATE":
        logger = setup_logging(PARAMS["log_dir"], PARAMS["log_base"])
        logger.info("Старт режима ГЕНЕРАЦИЯ")
        generate_all(logger)
    else:
        logger = setup_logging(PARAMS["log_dir"], PARAMS["log_base"])
        logger.error("Некорректный режим работы! Допустимы только 'GENERATE' или 'AGGREGATE'.")
    prog_end = time.time()
    logger.info(f"ВРЕМЯ РАБОТЫ ВСЕЙ ПРОГРАММЫ: {prog_end - prog_start:.2f} сек.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger = logging.getLogger()
        logger.error(f"Ошибка при выполнении: {e}\n{traceback.format_exc()}")
        print(f"Ошибка при выполнении: {e}")
        raise

